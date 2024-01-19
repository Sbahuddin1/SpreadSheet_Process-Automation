import openpyxl as xl
# from openpyxl.chart import BarChart3D, Reference


def process_workbook(filename, percentage):
    """
    :param filename: Takes the name of your file
    :param percentage: takes percentage ex: to lower price by
           10%, use 0.9 as percentage.
    :return: your file edited new file
    """

    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for i in range(2, sheet.max_row + 1):
        cl = sheet.cell(i, 3)
        corrected_price = cl.value * percentage
        corrected_price_cell = sheet.cell(i, 3)
        corrected_price_cell.value = corrected_price

    # TO add a 3D barchart of your new values:
    """    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row + 1,
                       min_col=3,
                       max_col=3)

    chart = BarChart3D()
    chart.add_data(values)
    sheet.add_chart(chart, 'f1')"""

    wb.save(filename)


process_workbook('transactions.xlsx', 0.9)
